from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.parser import extract_text
from app.crew_runner import run_agents

app = FastAPI()

UPLOAD_DIR = "uploads"
OUTPUT_DIR = "outputs"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/downloads", StaticFiles(directory=OUTPUT_DIR), name="downloads")


def _result_to_text(result) -> str:
    if hasattr(result, "raw") and result.raw:
        return result.raw
    return str(result)


def _download_path(file_path: str) -> str:
    return f"/downloads/{os.path.basename(file_path)}"


def _next_output_path(directory: str, base_name: str, extension: str) -> str:
    candidate = os.path.join(directory, f"{base_name}{extension}")
    if not os.path.exists(candidate):
        return candidate

    index = 2
    while True:
        candidate = os.path.join(directory, f"{base_name}{index}{extension}")
        if not os.path.exists(candidate):
            return candidate
        index += 1


def clean_json_text(raw_text: str) -> str:
    raw_text = raw_text.strip()

    if raw_text.startswith("```"):
        raw_text = re.sub(r"^```[a-zA-Z]*", "", raw_text).strip()
        raw_text = re.sub(r"```$", "", raw_text).strip()

    return raw_text


def extract_json_payload(raw_text: str) -> str:
    cleaned = clean_json_text(raw_text)

    # First try a fenced/clean full payload.
    if cleaned.startswith("[") and "]" in cleaned:
        return cleaned[: cleaned.rfind("]") + 1]

    if cleaned.startswith("{") and "}" in cleaned:
        return cleaned[: cleaned.rfind("}") + 1]

    array_start = cleaned.find("[")
    array_end = cleaned.rfind("]")
    if array_start != -1 and array_end != -1 and array_end > array_start:
        return cleaned[array_start : array_end + 1]

    object_start = cleaned.find("{")
    object_end = cleaned.rfind("}")
    if object_start != -1 and object_end != -1 and object_end > object_start:
        return cleaned[object_start : object_end + 1]

    return cleaned


def _extract_complete_objects(cleaned: str):
    start = cleaned.find("[")
    if start == -1:
        return None

    objects = []
    depth = 0
    in_string = False
    escape = False
    object_start = None

    for index in range(start, len(cleaned)):
        char = cleaned[index]

        if in_string:
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            continue

        if char == '"':
            in_string = True
            continue

        if char == "{":
            if depth == 0:
                object_start = index
            depth += 1
            continue

        if char == "}":
            if depth > 0:
                depth -= 1
                if depth == 0 and object_start is not None:
                    objects.append(cleaned[object_start:index + 1])
                    object_start = None

    parsed_objects = []
    for obj_text in objects:
        try:
            parsed_objects.append(json.loads(obj_text))
        except Exception:
            continue

    return parsed_objects or None


def parse_ai_output(raw_text: str):
    cleaned = extract_json_payload(raw_text)

    try:
        return json.loads(cleaned)
    except Exception:
        return _extract_complete_objects(cleaned)


def _pick(tc: dict, *keys: str) -> str:
    for key in keys:
        if key in tc and tc[key] is not None:
            return str(tc[key])
    return ""


def _compress_repeated_words(text: str, max_repeat: int = 3) -> str:
    words = text.split()
    if not words:
        return text

    compressed = []
    previous = None
    repeat_count = 0

    for word in words:
        normalized = word.lower()
        if normalized == previous:
            repeat_count += 1
        else:
            previous = normalized
            repeat_count = 1

        if repeat_count <= max_repeat:
            compressed.append(word)

    return " ".join(compressed)


def _sanitize_test_cases(test_cases):
    cleaned_cases = []
    field_limits = {
        "Test case No.": 40,
        "Function List/Test case description": 160,
        "Condition/Feature to be tested": 200,
        "steps": 500,
        "data set / values": 200,
        "expected_result": 260,
    }

    for case in test_cases:
        if not isinstance(case, dict):
            continue

        cleaned_case = {}
        for key, value in case.items():
            text = _compress_repeated_words("" if value is None else str(value))
            limit = field_limits.get(key)
            if limit and len(text) > limit:
                text = text[: limit - 3].rstrip() + "..."
            cleaned_case[key] = text

        cleaned_cases.append(cleaned_case)

    return cleaned_cases


def create_excel(test_cases, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    ws["A1"] = "Test Cases & Test Log"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="5B5B5B")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = "Test Info"
    ws["A2"].font = Font(bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor="7F6000")

    ws.merge_cells("A3:L3")
    ws["A3"] = "Source: Generated from uploaded FRS document"
    ws["A3"].fill = PatternFill("solid", fgColor="FFF2CC")

    ws.merge_cells("A4:L4")
    ws["A4"] = "Test Document: AI-generated structured test cases"
    ws["A4"].fill = PatternFill("solid", fgColor="FFF2CC")

    ws.merge_cells("A5:L5")
    ws["A5"] = "Remarks: Review generated test cases before final execution"
    ws["A5"].fill = PatternFill("solid", fgColor="FFF2CC")

    headers = [
        "Test Case No.",
        "Function List / Test case description",
        "Condition / Feature to be tested",
        "Data Set / Values",
        "Expected Result",
        "Actual Result",
        "Passed / Failed",
        "Testing Time (in Days)",
        "Man days / Default ID",
        "Defect Severity",
        "Defect Type",
        "Remarks"
    ]

    header_row = 7
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="C9A227")

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for cell in ws[header_row]:
        cell.border = thin_border

    start_row = header_row + 1

    for row_num, tc in enumerate(test_cases, start=start_row):
        row_data = [
            _pick(tc, "Test case No.", "Test Case No.", "Test case No"),
            _pick(tc, "Function List/Test case description", "Function List / Test case description"),
            _pick(tc, "Condition/Feature to be tested", "Condition / Feature to be tested"),
            _pick(tc, "data set / values", "Data Set / Values", "data_set", "dataset"),
            _pick(tc, "expected_result", "Expected Result", "expected result"),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    widths = {
        "A": 18,
        "B": 35,
        "C": 35,
        "D": 30,
        "E": 35,
        "F": 25,
        "G": 16,
        "H": 16,
        "I": 18,
        "J": 16,
        "K": 18,
        "L": 22,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[7].height = 36

    for row_num in range(start_row, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 55

    ws.freeze_panes = "A8"
    wb.save(output_path)


@app.post("/generate-testcases/")
async def generate_testcases(file: UploadFile = File(...)):
    file_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_path, "wb") as f:
        f.write(await file.read())

    frs_text = extract_text(file_path)
    result = run_agents(frs_text)
    raw_text = _result_to_text(result)

    text_output_path = _next_output_path(OUTPUT_DIR, "test_cases", ".txt")
    with open(text_output_path, "w", encoding="utf-8") as f:
        f.write(raw_text)

    with open(text_output_path, "r", encoding="utf-8") as f:
        saved_text = f.read()

    parsed_result = parse_ai_output(saved_text)
    if not parsed_result:
        return {
            "message": "TXT file generated, but AI output is not valid JSON so Excel could not be structured.",
            "txt_file": text_output_path,
            "txt_download": _download_path(text_output_path),
            "preview": saved_text[:1000]
        }

    if isinstance(parsed_result, dict):
        parsed_result = [parsed_result]

    parsed_result = _sanitize_test_cases(parsed_result)

    excel_output_path = _next_output_path(OUTPUT_DIR, "test_cases", ".xlsx")
    create_excel(parsed_result, excel_output_path)

    return {
        "message": "Success",
        "txt_file": text_output_path,
        "txt_download": _download_path(text_output_path),
        "excel_file": excel_output_path,
        "excel_download": _download_path(excel_output_path),
        "count": len(parsed_result)
    }
